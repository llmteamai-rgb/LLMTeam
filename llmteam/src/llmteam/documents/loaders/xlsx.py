"""
Excel loader for .xlsx files.

Uses openpyxl to load Excel workbooks.
"""

from pathlib import Path
from typing import Any, Dict, List, Union

from llmteam.documents.models import Document
from llmteam.documents.loaders.base import BaseLoader


class ExcelLoader(BaseLoader):
    """
    Loader for Excel files (.xlsx, .xls).

    Extracts text from all sheets, preserving table structure.

    Requires:
        pip install openpyxl
    """

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}

    def __init__(
        self,
        include_headers: bool = True,
        sheet_names: List[str] = None,
        include_sheet_name: bool = True,
    ):
        """
        Initialize the Excel loader.

        Args:
            include_headers: Include column headers in output.
            sheet_names: Specific sheets to load (None = all).
            include_sheet_name: Include sheet name in document metadata.
        """
        self.include_headers = include_headers
        self.sheet_names = sheet_names
        self.include_sheet_name = include_sheet_name

    def load(self, source: Union[str, Path]) -> List[Document]:
        """
        Load an Excel file.

        Args:
            source: Path to the Excel file.

        Returns:
            List of Documents, one per sheet.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "Excel support requires 'openpyxl'. "
                "Install with: pip install openpyxl"
            )

        path = Path(source)

        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        workbook = openpyxl.load_workbook(path, data_only=True)
        documents: List[Document] = []

        sheets_to_process = self.sheet_names or workbook.sheetnames

        for sheet_name in sheets_to_process:
            if sheet_name not in workbook.sheetnames:
                continue

            sheet = workbook[sheet_name]
            content = self._extract_sheet_content(sheet)

            if content.strip():
                metadata: Dict[str, Any] = {
                    "source": str(path),
                    "source_type": "excel",
                    "file_name": path.name,
                }

                if self.include_sheet_name:
                    metadata["sheet_name"] = sheet_name
                    metadata["sheet_index"] = workbook.sheetnames.index(sheet_name)

                documents.append(Document(content=content, metadata=metadata))

        workbook.close()
        return documents

    def _extract_sheet_content(self, sheet) -> str:
        """Extract text content from a sheet."""
        rows: List[List[str]] = []

        for row in sheet.iter_rows():
            row_values = []
            for cell in row:
                value = cell.value
                if value is not None:
                    row_values.append(str(value).strip())
                else:
                    row_values.append("")
            rows.append(row_values)

        # Remove empty rows from the end
        while rows and all(not v for v in rows[-1]):
            rows.pop()

        if not rows:
            return ""

        # Format as table text
        lines: List[str] = []

        if self.include_headers and rows:
            header_row = rows[0]
            lines.append(" | ".join(header_row))
            lines.append("-" * len(lines[0]))
            data_rows = rows[1:]
        else:
            data_rows = rows

        for row in data_rows:
            # Skip completely empty rows
            if any(v for v in row):
                lines.append(" | ".join(row))

        return "\n".join(lines)

    def supports(self, path: Union[str, Path]) -> bool:
        """Check if this loader supports the given file."""
        return self._get_extension(path) in self.SUPPORTED_EXTENSIONS


__all__ = ["ExcelLoader"]
